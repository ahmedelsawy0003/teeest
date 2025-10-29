"""معالجة ملفات إكسل الخاصة ببنود الكميات والتقارير."""
from __future__ import annotations

from io import BytesIO
from typing import Iterable, List

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font

from ..utils.constants import ERROR_MESSAGES
from ..utils.helpers import generate_code

EXPECTED_HEADERS = [
    "رقم البند",
    "اسم البند",
    "الوصف",
    "التصنيف",
    "الكمية",
    "الوحدة",
    "سعر الوحدة",
    "الإجمالي",
]


def parse_boq_excel(file_bytes: bytes) -> List[dict[str, str | float]]:
    """قراءة ملف إكسل والتحقق من صحة الأعمدة."""

    workbook = load_workbook(filename=BytesIO(file_bytes), data_only=True)
    sheet = workbook.active
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    if headers != EXPECTED_HEADERS:
        raise ValueError("ملف الإكسل لا يحتوي على الأعمدة المطلوبة")
    items: List[dict[str, str | float]] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        item = {
            "item_code": str(row[0]).strip(),
            "item_name": str(row[1]).strip(),
            "description": row[2] or "",
            "category": str(row[3]).strip(),
            "quantity": float(row[4] or 0),
            "unit": str(row[5]).strip(),
            "unit_price": float(row[6] or 0),
            "total_price": float(row[7] or 0),
        }
        items.append(item)
    return items


def export_boq_excel(project_name: str, items: Iterable[dict[str, str | float]]) -> bytes:
    """إنشاء ملف إكسل لبنود المشروع."""

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "بنود المشروع"
    sheet.sheet_view.rightToLeft = True
    header_font = Font(name="Cairo", bold=True)
    alignment = Alignment(horizontal="right")

    sheet.append(EXPECTED_HEADERS)
    for cell in sheet[1]:
        cell.font = header_font
        cell.alignment = alignment

    for item in items:
        sheet.append(
            [
                item.get("item_code"),
                item.get("item_name"),
                item.get("description"),
                item.get("category"),
                item.get("quantity"),
                item.get("unit"),
                item.get("unit_price"),
                item.get("total_price"),
            ]
        )

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
